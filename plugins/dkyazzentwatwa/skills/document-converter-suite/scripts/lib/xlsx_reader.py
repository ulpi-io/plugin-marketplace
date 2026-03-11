from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


@dataclass
class SheetContent:
    name: str
    cells: List[List[str]]  # 2D grid (rows x cols)


@dataclass
class XlsxContent:
    sheets: List[SheetContent]
    warnings: List[str] = field(default_factory=list)


def read_xlsx_content(path: Path, max_rows: int = 200, max_cols: int = 50) -> XlsxContent:
    """Read a bounded grid of values from each sheet with truncation warnings."""
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheets: List[SheetContent] = []
    warnings: List[str] = []

    for ws in wb.worksheets:
        # Get actual sheet dimensions
        actual_rows = ws.max_row
        actual_cols = ws.max_column

        # Check for truncation
        if actual_rows > max_rows:
            warnings.append(
                f"Sheet '{ws.title}': Truncated {actual_rows} rows → {max_rows} rows "
                f"({actual_rows - max_rows} rows omitted)"
            )

        if actual_cols > max_cols:
            warnings.append(
                f"Sheet '{ws.title}': Truncated {actual_cols} columns → {max_cols} columns "
                f"({actual_cols - max_cols} columns omitted)"
            )

        # Read grid with caps
        grid: List[List[str]] = []
        for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
            grid.append(["" if v is None else str(v) for v in row])
        sheets.append(SheetContent(name=ws.title, cells=grid))

    return XlsxContent(sheets=sheets, warnings=warnings)
