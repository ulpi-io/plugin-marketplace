from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook


def write_xlsx_from_sheets(output_path: Path, sheets: List[dict]) -> None:
    """Write an XLSX from a list of sheet dicts.

    Each sheet dict can have:
      - name: str
      - rows: list[list[str]]

    Notes:
      - This is a plain-value export (no formulas / styling).
    """
    wb = Workbook()

    # Remove default sheet if we are creating our own
    if sheets:
        default = wb.active
        wb.remove(default)

    for idx, sh in enumerate(sheets, start=1):
        name = (sh.get("name") or f"Sheet{idx}")[:31]
        ws = wb.create_sheet(title=name)
        rows = sh.get("rows") or []
        for r_i, row in enumerate(rows, start=1):
            for c_i, val in enumerate(row, start=1):
                ws.cell(row=r_i, column=c_i, value=val)

    if not sheets:
        wb.active.title = "Sheet1"

    wb.save(str(output_path))
