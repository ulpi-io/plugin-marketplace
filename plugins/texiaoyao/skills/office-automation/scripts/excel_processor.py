#!/usr/bin/env python3
"""
Excel 表格处理脚本
支持读取、写入、合并、转换、数据分析等功能
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
except ImportError:
    print("错误：请安装 openpyxl 和 pandas 库")
    print("运行：pip install openpyxl pandas")
    sys.exit(1)


def read_excel(file_path: str, sheet_name: str = None) -> dict:
    """读取 Excel 文件内容"""
    wb = load_workbook(file_path, data_only=True)
    
    result = {
        "file": file_path,
        "sheets": {},
        "sheet_names": wb.sheetnames
    }
    
    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames
    
    for sheet_name in sheets_to_read:
        if sheet_name not in wb.sheetnames:
            print(f"警告：工作表 '{sheet_name}' 不存在")
            continue
        
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        result["sheets"][sheet_name] = {
            "rows": len(data),
            "cols": len(data[0]) if data else 0,
            "data": data
        }
    
    return result


def write_excel(file_path: str, data: list, sheet_name: str = "Sheet1", headers: list = None):
    """写入数据到 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 写入表头
    if headers:
        ws.append(headers)
    
    # 写入数据
    for row in data:
        ws.append(row)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    print(f"✓ Excel 文件已保存：{file_path}")


def merge_excel_files(folder_path: str, output_path: str, sheet_name: str = "Merged"):
    """合并文件夹中的所有 Excel 文件"""
    folder = Path(folder_path)
    excel_files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xlsm"))
    
    if not excel_files:
        print(f"错误：在 {folder_path} 中未找到 Excel 文件")
        return
    
    all_data = []
    headers = None
    
    for file in excel_files:
        try:
            df = pd.read_excel(file)
            if headers is None:
                headers = df.columns.tolist()
            all_data.extend(df.values.tolist())
            print(f"✓ 已读取：{file.name}")
        except Exception as e:
            print(f"⚠ 跳过文件 {file.name}: {e}")
    
    if all_data:
        write_excel(output_path, all_data, sheet_name, headers)
        print(f"✓ 已合并 {len(excel_files)} 个文件，共 {len(all_data)} 行数据")
    else:
        print("错误：没有可合并的数据")


def convert_excel(file_path: str, to_format: str = "csv", output_path: str = None):
    """转换 Excel 文件格式"""
    if output_path is None:
        output_path = Path(file_path).with_suffix(f".{to_format}")
    
    if to_format.lower() == "csv":
        # Excel 转 CSV
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            csv_path = Path(output_path).parent / f"{Path(output_path).stem}_{sheet_name}.csv"
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"✓ 已导出：{csv_path}")
    
    elif to_format.lower() == "xlsx":
        # CSV 转 Excel
        df = pd.read_csv(file_path, encoding='utf-8')
        df.to_excel(output_path, index=False)
        print(f"✓ 已转换：{output_path}")
    
    else:
        print(f"错误：不支持的格式 '{to_format}'")


def analyze_excel(file_path: str, sheet_name: str = None, pivot: bool = False):
    """分析 Excel 数据"""
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    print(f"\n📊 数据分析报告：{Path(file_path).name}")
    print("=" * 50)
    print(f"行数：{len(df)}")
    print(f"列数：{len(df.columns)}")
    print(f"\n列名：{list(df.columns)}")
    
    print(f"\n数据类型：")
    for col, dtype in df.dtypes.items():
        print(f"  {col}: {dtype}")
    
    print(f"\n缺失值统计：")
    missing = df.isnull().sum()
    for col, count in missing.items():
        if count > 0:
            print(f"  {col}: {count} ({count/len(df)*100:.1f}%)")
    
    if not missing.any():
        print("  无缺失值 ✓")
    
    print(f"\n数值列统计：")
    numeric_cols = df.select_dtypes(include=['number']).columns
    if len(numeric_cols) > 0:
        print(df[numeric_cols].describe().to_string())
    else:
        print("  无数值列")
    
    if pivot:
        print(f"\n⚠️  透视表需要提供行列参数，使用 --pivot-rows 和 --pivot-cols 指定")


def main():
    parser = argparse.ArgumentParser(description="Excel 表格处理工具")
    subparsers = parser.add_subparsers(dest="command", help="命令")
    
    # read 命令
    read_parser = subparsers.add_parser("read", help="读取 Excel 内容")
    read_parser.add_argument("file", help="Excel 文件路径")
    read_parser.add_argument("--sheet", help="工作表名称（不指定则读取全部）")
    
    # write 命令
    write_parser = subparsers.add_parser("write", help="创建 Excel 文件")
    write_parser.add_argument("output", help="输出文件路径")
    write_parser.add_argument("--data", required=True, help="JSON 格式数据")
    write_parser.add_argument("--sheet", default="Sheet1", help="工作表名称")
    write_parser.add_argument("--headers", help="表头（逗号分隔）")
    
    # merge 命令
    merge_parser = subparsers.add_parser("merge", help="合并 Excel 文件")
    merge_parser.add_argument("folder", help="包含 Excel 文件的文件夹")
    merge_parser.add_argument("--output", required=True, help="输出文件路径")
    merge_parser.add_argument("--sheet", default="Merged", help="工作表名称")
    
    # convert 命令
    convert_parser = subparsers.add_parser("convert", help="转换文件格式")
    convert_parser.add_argument("file", help="输入文件路径")
    convert_parser.add_argument("--to", required=True, choices=["csv", "xlsx"], help="目标格式")
    convert_parser.add_argument("--output", help="输出文件路径")
    
    # analyze 命令
    analyze_parser = subparsers.add_parser("analyze", help="数据分析")
    analyze_parser.add_argument("file", help="Excel 文件路径")
    analyze_parser.add_argument("--sheet", help="工作表名称")
    analyze_parser.add_argument("--pivot", action="store_true", help="生成透视表")
    
    args = parser.parse_args()
    
    if args.command == "read":
        result = read_excel(args.file, args.sheet)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    
    elif args.command == "write":
        data = json.loads(args.data)
        headers = args.headers.split(',') if args.headers else None
        write_excel(args.output, data, args.sheet, headers)
    
    elif args.command == "merge":
        merge_excel_files(args.folder, args.output, args.sheet)
    
    elif args.command == "convert":
        convert_excel(args.file, args.to, args.output)
    
    elif args.command == "analyze":
        analyze_excel(args.file, args.sheet, args.pivot)
    
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
